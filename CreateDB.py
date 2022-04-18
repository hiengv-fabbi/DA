# Tìm đặc trưng của tập dữ liệu ảnh
# vì đã tìm đặc trưng của tập ảnh trong HoaQua rồi nên 0 cần chạy lại lần 2
from Function import feat,Average
import cv2
import numpy as np
import os
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
er = 0
ec = 1
files = []


# Tạo file excel name lưu trữ tên và đặc trưng của ảnh dữ liệu 
path = 'Corel-100'
wbname = xlsxwriter.Workbook('Corel-100-name.xlsx')
wsname = wbname.add_worksheet('name')



# Duyệt các ảnh trong file
for r, d, f in os.walk(path):
    for file in f:
        if '.jpg' in file:
            files.append(os.path.join("D:/DA/",r, file))


# Lưu tên ảnh vào file name
for f in files:
    wsname.write(er,0,f)
# thuc hien tinh feature cua anh
    print(feat(f))
    #lưu các giá trị tính được vào file excel
    for i in (feat(f)):
        wsname.write(er, ec,i)
        ec += 1
    # ws.cell(er, 6).value = f\
    ec = 1
    er +=1

wbname.close()

