import cv2
import numpy as np
import os
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook

# Tạo file excel distance lưu trữ khoảng cách euclide của các ảnh dữ liệu
wbdis = xlsxwriter.Workbook('Corel-100-distance.xlsx')
wsdis = wbdis.add_worksheet('distance')

er = 0
ec = 0
wb = load_workbook("Corel-100-name.xlsx")
ws = wb.active
# Lưu khoảng cách euclide vào file distance 
for row in ws.iter_rows(min_col=2, max_col=1001, values_only=True):
    feat_data = np.array(row)
    for row1 in ws.iter_rows(min_col=2, max_col=1001, values_only=True):
        feat_data1 = np.array(row1) 
        dist = np.linalg.norm(feat_data - feat_data1)
        wsdis.write(er,ec,dist)
        print(ec)
        ec+=1
    ec = 0
    er += 1



# wb.save('fruits-name.xlsx')
wbdis.close()