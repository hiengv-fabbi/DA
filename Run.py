from Function import feat,Average
import cv2
import numpy as np
import os 
import pathlib
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from PIL import Image,ImageTk
import matplotlib.image as mpimg
import networkx as nx
from networkx.algorithms import tree
from tkinter import *
from PIL import Image, ImageTk
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)


root = Tk()


def findIMG():
    #Duyệt file name và file distance 
    wb = load_workbook("Corel-100-name.xlsx")
    ws = wb.active 
    wbdis = load_workbook("Corel-100-distance.xlsx")
    wsdis = wbdis.active 

    # Nhập ảnh cần truy xuất
    # query =input("Enter your value: ")
    global img1, photo 

    query = filedialog.askopenfilename()
    

    img1 = ImageTk.PhotoImage(Image.open(query))
    photo.configure(image=img1)

    feat_query = np.array(feat(query))

    ht_query = os.path.split(query)

    # Lưu tên ảnh và đặc trưng của ảnh vào file name
    ec = 2
    ws.cell(101, 1).value = query
    for i in feat_query:
        ws.cell(101, ec).value = i
        ec += 1

    wb.save('Corel-100-name.xlsx')

    r = 1
    c = 1
    name = []
    distance = []


    qr = 101
    qc = 101

    # Duyệt các đặc trưng của ảnh trong file distance và tính khoảng cách với ảnh truy vấn
    # Lưu khoảng cách của ảnh truy vấn với tập ảnh dữ liệu vào file distance
    for row in ws.iter_rows(min_col=2, max_col=1001, values_only=True):
        feat_data = np.array(row)
        dist = np.linalg.norm(feat_data - feat_query)
        wsdis.cell(r, qc).value = dist
        wsdis.cell(qr, c).value = dist
        r+=1
        c+=1

    wbdis.save('Corel-100-distance.xlsx')

    # Duyệt tên ảnh vào mảng name
    for row in ws.iter_rows(min_col=0, max_col=1, values_only=True):
        name.append(row)
    # Duyệt khoảng cách ảnh vào mảng distance
    for row in wsdis.iter_rows(min_col=0, max_col=101, values_only=True):
        distance.append(row)


    # tạo đồ thị đày đủ G
    G = nx.Graph()
    for store in name:
        G.add_node(store)

    for i in range(101):
        for j in range(101):
            if name[i]!=name[j]:
                G.add_edge(name[i],name[j],weight=distance[i][j])

    # áp dụng thuật toán krusal để tạo cây khung nhỏ nhất G1
    mst = tree.minimum_spanning_edges(G, algorithm="kruskal", data=True)      
    
    edgelist = list(mst)

    G1 = nx.Graph()
    G1.add_edges_from(edgelist)

    # plt.figure(num=None, figsize=(10, 10), dpi=1200)
    color_map = []
    for node in G1:
        if node == (query,):
            color_map.append('blue')
        else: 
            color_map.append('green')      


    


    listweight = [attrs["weight"] for a, b, attrs in G1.edges(data=True)]
    Average(listweight)


    edgerm = [(a,b,attrs) for a, b, attrs in G1.edges(data=True) if attrs["weight"] > Average(listweight)]


    G1.remove_edges_from(edgerm)

    fig2 = plt.figure()
    nx.draw(G1,node_color=color_map)

    img = nx.node_connected_component(G1, (query,))

    count = 0
    fig = plt.figure(figsize=(7, 7))
    pos = 1

    for i in img: 
        str = ''.join(i)
        # print(str)
        ht_data = os.path.split(str)
        # print(ht_data[0], " ?= " , ht_query[0])
        if (ht_data[0] == ht_query[0]):
            count +=1
        image = mpimg.imread(str)
        fig.add_subplot(7, 7, pos)
        pos += 1
        plt.imshow(image)
        plt.axis('off')
       
    # print(img)
    # plt.show()

    canvas = FigureCanvasTkAgg(fig, master = root)
    canvas.get_tk_widget().place(x = 400, y=0 , width =1000, height =800)
    canvas.draw()

    canvas2 = FigureCanvasTkAgg(fig2, master = root)
    canvas2.get_tk_widget().place(x = 0, y= 260 , width =400, height =560)
    canvas2.draw()


    # toolbar = NavigationToolbar2Tk(canvas2,root)
    # toolbar.update()
    # (count/len(img))*100, '%'
    print('Độ chính xác: ', count , "/" , len(img) ," * 100 % = ", (count/len(img))*100, '%')



photo = Label(root)
photo.place(x=20,y=20)


button = Button(root, text="Tìm kiếm",command=findIMG)
button.place(x=50, y=125)

label = Label(root, text="Chọn ảnh cần tìm kiếm:")
label.place(x=20, y=100)


root.geometry("1400x800+10+20")
# root.eval('tk::PlaceWindow . center')
root.mainloop()